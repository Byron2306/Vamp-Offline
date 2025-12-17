from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from backend.staff_profile import DEFAULT_KPAS, KPA, StaffProfile, staff_is_director_level


DEFAULT_KPA_NAMES: Dict[str, str] = {
    "KPA1": "Teaching and Learning, including Higher Degree Supervision",
    "KPA2": "Personal Research, Innovation and/or Creative Outputs",
    "KPA3": "Social Responsiveness and Industry Involvement",
    "KPA4": "Academic Leadership, Management and Administration",
    "KPA5": "OHS (Occupational Health and Safety)",
    "KPA6": "People Management",
}

PA_ORDER: List[Tuple[str, str]] = [
    ("KPA1", DEFAULT_KPA_NAMES["KPA1"]),
    ("KPA2", DEFAULT_KPA_NAMES["KPA2"]),
    ("KPA3", DEFAULT_KPA_NAMES["KPA3"]),
    ("KPA4", DEFAULT_KPA_NAMES["KPA4"]),
    ("KPA5", DEFAULT_KPA_NAMES["KPA5"]),
]

HEADERS: List[str] = ["KPA Name", "Outputs", "KPIs", "Weight", "Hours", "Outcomes", "Active"]

TASK_KEYS = {
    "teaching",
    "supervision",
    "research",
    "leadership",
    "social",
    "ohs",
    "teaching_modules",
    "teaching_practice_windows",
    "people_management",
    "people_management_items",
}

SKIP_KEYS = {"hours", "weight_pct", "name", "norm_hours", "kpa_summary"}


def _normalise_name(name: str) -> str:
    return " ".join(name.lower().split())


def _kpa_lookup(profile: StaffProfile) -> Dict[str, KPA]:
    lookup: Dict[str, KPA] = {}
    for kpa in profile.kpas:
        lookup[kpa.code.upper()] = kpa
        lookup[_normalise_name(kpa.name)] = kpa
    return lookup


def _safe_float(value: Any) -> float:
    try:
        f = float(value)
    except Exception:
        return 0.0
    return 0.0 if f != f else f


def _coerce_lines_from(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        return [text] if text else []
    if isinstance(value, dict):
        lines: List[str] = []
        for sub in value.values():
            lines.extend(_coerce_lines_from(sub))
        return lines
    if isinstance(value, Iterable):
        lines = []
        for item in value:
            lines.extend(_coerce_lines_from(item))
        return lines
    return []


def _outcomes_from_ta_context(context: Dict[str, Any] | None) -> str:
    if not isinstance(context, dict):
        return ""

    ordered_lines: List[str] = []
    seen: set[str] = set()

    for key, value in context.items():
        if key in SKIP_KEYS:
            continue
        if key not in TASK_KEYS and not isinstance(value, (list, tuple, set, dict, str)):
            continue

        lines = _coerce_lines_from(value)
        for line in lines:
            if not line or line in seen:
                continue
            seen.add(line)
            ordered_lines.append(line)

    return "\n".join(f"• {line}" for line in ordered_lines)


def _apply_layout(ws) -> None:
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 8

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=3, max_col=7):
        for cell in row:
            if cell.column_letter in {"B", "C", "F"}:
                cell.alignment = wrap_alignment

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    for cell in ws[2]:
        cell.font = bold_font


def _resolve_kpa_name(kpa: KPA | None, code: str) -> str:
    if kpa and kpa.name:
        return kpa.name
    for default_code, default_name in DEFAULT_KPAS:
        if default_code.upper() == code.upper():
            return DEFAULT_KPA_NAMES.get(default_code, default_name)
    return DEFAULT_KPA_NAMES.get(code.upper(), code)


def generate_pa_skeleton_from_ta(profile: StaffProfile, out_dir: Path) -> Tuple[Path, List[List[object]]]:
    lookup = _kpa_lookup(profile)

    ordered_codes = list(PA_ORDER)
    if staff_is_director_level(profile):
        ordered_codes.append(("KPA6", DEFAULT_KPA_NAMES["KPA6"]))

    rows: List[List[object]] = []
    for code, fallback_name in ordered_codes:
        kpa = lookup.get(code) or lookup.get(_normalise_name(fallback_name))
        kpa_name = _resolve_kpa_name(kpa, code) if kpa else fallback_name

        ta_context = kpa.ta_context if kpa else {}
        weight = _safe_float((ta_context or {}).get("weight_pct") or (kpa.weight if kpa else 0.0))
        hours = _safe_float((ta_context or {}).get("hours") or (kpa.hours if kpa else 0.0))

        # HARD LOCK: OHS (KPA5) must always weigh 2% and must not include module codes.
        if code == "KPA5":
            weight = 2.0


        outcomes = _outcomes_from_ta_context(ta_context)
        if code == "KPA5":
            outcomes = "Compliance with institutional Occupational Health and Safety requirements"

        if not outcomes:
            outcomes = "Task Agreement tasks to be detailed"

        active_flag = "Y"

        rows.append([kpa_name, "", "", weight, hours, outcomes, active_flag])

    wb = Workbook()
    ws = wb.active
    ws.title = "pa-report"

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Performance Agreement {profile.staff_id} {profile.cycle_year}"
    title_cell.font = Font(bold=True)

    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    _apply_layout(ws)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"PA_{profile.staff_id}_{profile.cycle_year}_skeleton.xlsx"
    wb.save(out_path)

    return out_path, rows
