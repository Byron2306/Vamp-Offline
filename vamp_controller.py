from __future__ import annotations

"""
VampController (Offline)
------------------------
This controller is designed to be *robust across historic UI/API versions*.

Primary fixes included:
- Stable profile enrolment (enrol_staff)
- TA import that correctly detects the TA's percent column ("%" header) and weighted rows
- Build expectations via expectation_engine.build_expectations_from_ta with flexible signature
- Persist tasks into progress_tracker (SQLite) without KeyErrors
- Provide minimal final-PA generator compatible with your existing UI expectations

NOTE:
- Scanner/AI features are intentionally stubbed here (return ok + placeholders) to keep core pipeline stable.
"""

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json
import re
import uuid

import openpyxl

from backend.expectation_engine import build_expectations_from_ta
from backend import progress_tracker


PROJECT_ROOT = Path(__file__).resolve().parents[2]  # backend/services -> project root
OFFLINE_RESULTS_DIR = PROJECT_ROOT / "output" / "offline_results"
OFFLINE_RESULTS_DIR.mkdir(parents=True, exist_ok=True)


def _safe_int(v: Any, default: int = 0) -> int:
    try:
        if v is None:
            return default
        s = str(v).strip()
        if not s or s.lower() in {"unknown", "n/a", "na", "none", "null"}:
            return default
        return int(float(s))
    except Exception:
        return default


def _safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace("%", "")
        if not s or s.lower() in {"unknown", "n/a", "na", "none", "null"}:
            return default
        return float(s)
    except Exception:
        return default


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _clean_text(v: Any) -> str:
    s = _s(v)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _kpa_normalise(name: str) -> Tuple[str, str]:
    """
    Map common KPA labels to the standard NWU five-KPA set.
    """
    n = (name or "").lower()
    if "ohs" in n or "health" in n or "safety" in n:
        return ("KPA5", "Occupational Health and Safety")
    if "social" in n or "community" in n or "industry" in n or "engagement" in n or "respons" in n:
        return ("KPA4", "Social Responsiveness / Community and Industry Engagement")
    if "lead" in n or "management" in n or "admin" in n or "govern" in n:
        return ("KPA3", "Academic Leadership and Management")
    if "teach" in n or "learning" in n or "assessment" in n or "efundi" in n:
        return ("KPA1", "Teaching and Learning")
    if "research" in n or "innovation" in n or "creative" in n or "publication" in n or "supervis" in n:
        return ("KPA2", "Research and Innovation / Creative Outputs")
    return ("KPA1", "Teaching and Learning")


def _parse_section_cell(a: Any, b: Any) -> Tuple[str, str]:
    """
    'SECTION 1\n...title...' often sits in column A (merged/linebreak).
    Sometimes title is in column B.
    """
    a_txt = _s(a)
    b_txt = _s(b)
    if not a_txt:
        return ("", b_txt)

    if a_txt.upper().startswith("SECTION"):
        parts = [p.strip() for p in a_txt.splitlines() if p.strip()]
        sec_id = parts[0] if parts else a_txt.strip()
        title = ""
        if len(parts) > 1:
            title = " ".join(parts[1:])
        if not title:
            title = b_txt
        return (sec_id, title)

    return ("", "")


def parse_task_agreement_xlsx(path: str) -> Dict[str, Any]:
    """
    Parse the NWU TA template you uploaded (where weighting header is literally '%').

    Returns:
      {
        "ok": True,
        "sheet": "...",
        "tasks": [
          {"section_id":"SECTION 1", "section_title":"...", "output":"...", "hours": 120, "weight_pct": 4.0},
          ...
        ],
        "diagnostics": {...}
      }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    # Prefer the explicit sheet name if present
    sheet = "Task Agreement Form" if "Task Agreement Form" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    # Detect header row by finding "NUMBER OF HOURS" and a '%' column
    header_row = None
    col_hours = None
    col_pct = None
    col_desc = None
    col_item = None

    # scan first 80 rows for header candidates
    for r in range(1, min(ws.max_row, 120) + 1):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 40) + 1)]
        row_str = [(_s(v).lower()) for v in row]
        if any("number of hours" in s for s in row_str) and any(s in {"%", "percent", "percentage"} or "%" in s for s in row_str):
            header_row = r
            # locate columns
            for c, v in enumerate(row, start=1):
                sv = _s(v).strip().lower()
                if "number of hours" in sv:
                    col_hours = c
                if sv == "%" or "percent" in sv or "percentage" in sv or sv.endswith("%"):
                    col_pct = c
                if "task agreement kpa" in sv or "calc" in sv or "task agreement" in sv:
                    col_desc = c
                if sv in {"item", "items"}:
                    col_item = c
            break

    # Fallbacks based on the known template layout
    col_item = col_item or 1
    col_desc = col_desc or 2
    col_hours = col_hours or 4
    col_pct = col_pct or 5

    if header_row is None:
        # Still try to parse using these defaults, but record diagnostics
        header_row = 1

    tasks: List[Dict[str, Any]] = []
    current_section_id = ""
    current_section_title = ""

    weighted_rows = 0
    considered_rows = 0

    for r in range(header_row + 1, ws.max_row + 1):
        item = ws.cell(r, col_item).value
        desc = ws.cell(r, col_desc).value
        hours = ws.cell(r, col_hours).value
        pct = ws.cell(r, col_pct).value

        # detect new section
        sec_id, sec_title = _parse_section_cell(item, desc)
        if sec_id:
            current_section_id = sec_id
            current_section_title = sec_title
            continue

        # stop if we hit end markers (template has long tail, but we only care about main table)
        if isinstance(item, str) and item.strip().upper().startswith("GRAND TOTAL"):
            break

        d_txt = _clean_text(desc)
        if not d_txt:
            continue

        considered_rows += 1
        p = _safe_float(pct, 0.0)
        h = _safe_float(hours, 0.0)

        # skip obvious totals/subtotals
        up = d_txt.upper()
        if "TOTAL" in up and len(d_txt) <= 40:
            continue

        if p > 0:
            weighted_rows += 1

        # only tasks with a weighting count as outputs for the PA/expectations
        if p <= 0:
            continue

        tasks.append(
            {
                "section_id": current_section_id,
                "section_title": current_section_title,
                "output": d_txt,
                "hours": h,
                "weight_pct": p,
            }
        )

    return {
        "ok": True,
        "sheet": sheet,
        "tasks": tasks,
        "diagnostics": {
            "header_row": header_row,
            "col_item": col_item,
            "col_desc": col_desc,
            "col_hours": col_hours,
            "col_pct": col_pct,
            "rows_considered": considered_rows,
            "weighted_rows_seen": weighted_rows,
            "tasks_emitted": len(tasks),
        },
    }


class VampController:
    def __init__(self, output_dir: Optional[str | Path] = None) -> None:
        self.output_dir = Path(output_dir) if output_dir else OFFLINE_RESULTS_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
        # Ensure DB exists
        progress_tracker.init_db()

    # -----------------------
    # Profile
    # -----------------------

    def enrol_staff(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """
        Create/load a profile stub. We keep it simple: just persist to a JSON file.
        """
        payload = dict(payload or {})
        staff_id = _s(payload.get("staff_id") or payload.get("staffId") or payload.get("id"))
        year = _safe_int(payload.get("year") or payload.get("cycle_year"), date.today().year)

        if not staff_id:
            return {"ok": False, "error": "staff_id is required"}

        profile = {
            "staff_id": staff_id,
            "year": year,
            "name": _s(payload.get("name") or payload.get("full_name") or ""),
            "campus": _s(payload.get("campus") or ""),
            "unit": _s(payload.get("unit") or payload.get("school") or ""),
            "created_at": datetime.utcnow().isoformat(timespec="seconds"),
        }
        p = self.output_dir / f"profile_{staff_id}_{year}.json"
        p.write_text(json.dumps(profile, indent=2, ensure_ascii=False), encoding="utf-8")
        return {"ok": True, "profile": profile, "profile_path": str(p)}

    # compatibility aliases used by older api.py versions
    enrol_profile = enrol_staff
    enrol_or_load_profile = enrol_staff
    enrol_or_load = enrol_staff
    create_or_load_profile = enrol_staff

    # -----------------------
    # TA Import
    # -----------------------

    def import_task_agreement(self, staff_id: str, year: int, ta_path: str) -> Dict[str, Any]:
        sid = _s(staff_id)
        y = _safe_int(year, date.today().year)

        ta_summary = parse_task_agreement_xlsx(ta_path)
        ta_summary["staff_id"] = sid
        ta_summary["year"] = y

        # Fail fast with diagnostics if we found no weighted tasks
        if not ta_summary.get("tasks"):
            diag = ta_summary.get("diagnostics") or {}
            raise RuntimeError(
                "TA parsed but no weighted tasks were found. "
                "This usually means the parser did not locate the '%' (weighting) column "
                f"or the template differs. Diagnostics: {diag}"
            )

        # Build expectations from the TA summary (signature-flexible)
        expectations = build_expectations_from_ta(sid, y, ta_summary)

        # Persist artifacts
        imports_dir = self.output_dir / "imports"
        imports_dir.mkdir(parents=True, exist_ok=True)

        ta_out = self.output_dir / f"ta_summary_{sid}_{y}.json"
        ta_out.write_text(json.dumps(ta_summary, indent=2, ensure_ascii=False), encoding="utf-8")

        exp_out = self.output_dir / f"expectations_{sid}_{y}.json"
        exp_out.write_text(json.dumps(expectations, indent=2, ensure_ascii=False), encoding="utf-8")

        # Upsert tasks into SQLite
        progress_tracker.upsert_tasks(expectations.get("tasks") or [])

        return {"ok": True, "ta_summary": ta_summary, "expectations": expectations, "ta_path": ta_path}

    # alias
    import_ta = import_task_agreement

    def rebuild_expectations(self, staff_id: str, year: int) -> Dict[str, Any]:
        sid = _s(staff_id)
        y = _safe_int(year, date.today().year)
        ta_out = self.output_dir / f"ta_summary_{sid}_{y}.json"
        if not ta_out.exists():
            return {"ok": False, "error": "No TA summary found. Import a TA first."}
        ta_summary = json.loads(ta_out.read_text(encoding="utf-8"))
        expectations = build_expectations_from_ta(sid, y, ta_summary)
        (self.output_dir / f"expectations_{sid}_{y}.json").write_text(
            json.dumps(expectations, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        progress_tracker.upsert_tasks(expectations.get("tasks") or [])
        return expectations

    rebuild = rebuild_expectations

    def load_expectations(self, staff_id: str, year: int) -> Dict[str, Any]:
        sid = _s(staff_id)
        y = _safe_int(year, date.today().year)
        p = self.output_dir / f"expectations_{sid}_{y}.json"
        if p.exists():
            return json.loads(p.read_text(encoding="utf-8"))
        # fallback: read from DB only
        tasks = progress_tracker.list_tasks(sid, y)
        return {"ok": True, "staff_id": sid, "year": y, "tasks": tasks, "months": [f"{y}-{m:02d}" for m in range(1,13)], "by_month": {}}

    get_expectations = load_expectations

    # -----------------------
    # Progress + Evidence
    # -----------------------

    def progress_summary(self, staff_id: str, year: int, month_bucket: str) -> Dict[str, Any]:
        return progress_tracker.get_progress_summary(_s(staff_id), _safe_int(year, date.today().year), _s(month_bucket))

    get_progress_summary = progress_summary
    progress = progress_summary

    def list_evidence(self, staff_id: str, year: int, limit: int = 200) -> Dict[str, Any]:
        return progress_tracker.list_evidence(_s(staff_id), _safe_int(year, date.today().year), limit)

    evidence_list = list_evidence

    # -----------------------
    # Reports
    # -----------------------

    def generate_final_pa(self, staff_id: str, year: int) -> Dict[str, Any]:
        """
        Minimal PA generator:
        - Groups tasks by KPA (using kpa_name already assigned by expectations engine)
        - Writes an Excel file similar to your 'pa-report' sheet layout (core columns only)
        """
        sid = _s(staff_id)
        y = _safe_int(year, date.today().year)

        exp = self.load_expectations(sid, y)
        tasks = exp.get("tasks") or []
        if not tasks:
            return {"ok": False, "error": "No tasks found. Import TA first."}

        # group
        by_kpa: Dict[str, Dict[str, Any]] = {}
        for t in tasks:
            code = _s(t.get("kpa_code") or "")
            name = _s(t.get("kpa_name") or "")
            if not name:
                code, name = _kpa_normalise(_s(t.get("section_title") or ""))

            if code not in by_kpa:
                by_kpa[code] = {"kpa_name": name, "outputs": [], "weight": 0.0, "hours": 0.0}
            by_kpa[code]["outputs"].append(_s(t.get("output") or ""))
            by_kpa[code]["weight"] += _safe_float(t.get("weight_pct"), 0.0)
            by_kpa[code]["hours"] += _safe_float(t.get("hours"), 0.0)

        # order KPAs
        order = ["KPA1", "KPA2", "KPA4", "KPA3", "KPA5"]
        ordered = [(k, by_kpa[k]) for k in order if k in by_kpa] + [(k, v) for k, v in by_kpa.items() if k not in order]

        # write xlsx
        out = self.output_dir / f"PA_{sid}_{y}_final.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "pa-report"

        ws.cell(1, 1).value = f"Performance Agreement {sid} {y}"
        headers = ["KPA Name", "Outputs", "KPIs", "Weight", "Hours", "Outcomes"]
        for c, h in enumerate(headers, start=1):
            ws.cell(2, c).value = h

        row = 3
        for _, info in ordered:
            ws.cell(row, 1).value = info["kpa_name"]
            ws.cell(row, 2).value = "\n".join([f"• {o}" for o in info["outputs"] if o])
            ws.cell(row, 3).value = ""  # KPIs optional (can be mapped later)
            ws.cell(row, 4).value = round(float(info["weight"]), 3)
            ws.cell(row, 5).value = round(float(info["hours"]), 2)
            ws.cell(row, 6).value = ""  # Outcomes optional (can be mapped later)
            row += 1

        wb.save(out)
        return {"ok": True, "path": str(out)}

    final_pa = generate_final_pa
    generate_pa = generate_final_pa

    # -----------------------
    # Scanner placeholders
    # -----------------------

    def start_scan(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        return {"ok": True, "run_id": str(uuid.uuid4()), "note": "Scanner not implemented in this controller build."}

    def stop_scan(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        return {"ok": True}

    def scan_events(self, *args: Any) -> Dict[str, Any]:
        return {"ok": True, "events": [], "after": 0}

    get_scan_events = scan_events
