from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
from openpyxl import Workbook

from backend.staff_profile import StaffProfile

REPO_ROOT = Path(__file__).resolve().parents[2]
EVIDENCE_DIR = REPO_ROOT / "backend" / "data" / "evidence"

KPA_DEFINITIONS: Dict[str, str] = {
    "KPA1": "Teaching and Learning",
    "KPA2": "Occupational Health and Safety",
    "KPA3": "Research and Innovation / Creative Outputs",
    "KPA4": "Academic Leadership and Management",
    "KPA5": "Social Responsiveness / Community and Industry Engagement",
}


def _render_kpi_cell(kpi: Any) -> str:
    """Combine KPI description with measurable details for export."""

    description = (getattr(kpi, "description", "") or "").strip()
    details: List[str] = []

    measure = (getattr(kpi, "measure", "") or "").strip()
    target = (getattr(kpi, "target", "") or "").strip()
    due = (getattr(kpi, "due", "") or "").strip()
    evidence_types = getattr(kpi, "evidence_types", []) or []

    if measure:
        details.append(f"Measure: {measure}")
    if target:
        details.append(f"Target: {target}")
    if due:
        details.append(f"Due: {due}")
    evidence_text = "; ".join(str(e).strip() for e in evidence_types if str(e).strip())
    if evidence_text:
        details.append(f"Evidence: {evidence_text}")

    if details:
        suffix = "; ".join(details)
        return f"{description} ({suffix})" if description else suffix

    return description


def _rows_for_profile(profile: StaffProfile) -> List[List[Any]]:
    """
    Flatten profile KPAs/KPIs into rows:

    [KPA Name, Outputs, KPI, Weight, Hours, Outcomes, Active]
    """
    rows: List[List[Any]] = []
    for kpa in profile.kpas:
        kpa_name = kpa.name or KPA_DEFINITIONS.get(kpa.code, kpa.code)
        if not kpa.kpis:
            rows.append([kpa_name, "", "", kpa.weight, kpa.hours, "", "Y"])
            continue

        for kpi in kpa.kpis:
            kpi_cell = _render_kpi_cell(kpi)
            rows.append(
                [
                    kpa_name,
                    (kpi.outputs or "").strip(),
                    kpi_cell,
                    kpi.weight if kpi.weight is not None else 0.0,
                    kpi.hours if kpi.hours is not None else 0.0,
                    (kpi.outcomes or "").strip(),
                    "Y" if getattr(kpi, "active", True) else "N",
                ]
            )
    return rows


def _filter_live_kpis(rows: List[List[Any]]) -> List[List[Any]]:
    """
    Filter rows to include only 'live' KPIs with measurable allocations.

    Rules:
      - Weight > 0 OR Hours > 0
      - Active flag in column 7 is 'Y' / 'YES' / 'TRUE' (case-insensitive)

    Row layout: [KPA Name, Outputs, KPI, Weight, Hours, Outcomes, Active]
    """
    filtered: List[List[Any]] = []
    for row in rows:
        if len(row) < 7:
            continue
        try:
            weight = float(row[3] or 0)
        except Exception:
            weight = 0.0
        try:
            hours = float(row[4] or 0)
        except Exception:
            hours = 0.0
        active_str = str(row[6]).strip().upper()
        if (weight > 0 or hours > 0) and active_str in {"Y", "YES", "TRUE"}:
            filtered.append(row)
    return filtered


def _load_evidence(staff_id: str, cycle_year: int) -> pd.DataFrame:
    """
    Load evidence CSV for a staff/year pair.

    Expected filename: evidence_{staff_id}_{cycle_year}.csv
    Returns empty DataFrame if not found.
    """
    path = EVIDENCE_DIR / f"evidence_{staff_id}_{cycle_year}.csv"
    if not path.exists():
        return pd.DataFrame()

    df = pd.read_csv(path, encoding="utf-8")
    if "month_bucket" not in df.columns:
        return df

    # Extract month as integer from YYYY-MM
    try:
        df["month_bucket"] = df["month_bucket"].astype(str)
        df["month_int"] = df["month_bucket"].str[-2:].astype(int)
    except Exception:
        df["month_int"] = 0
    return df


def _aggregate_kpa_window(df: pd.DataFrame, month_min: int, month_max: int) -> pd.DataFrame:
    """
    Aggregate evidence in a month window [month_min, month_max] by KPA.
    """
    if df.empty:
        return pd.DataFrame(
            columns=[
                "kpa_code",
                "kpa_name",
                "count_evidence",
                "avg_rating",
                "max_rating",
                "ratings_labels",
                "sample_impact",
            ]
        )

    window = df[(df["month_int"] >= month_min) & (df["month_int"] <= month_max)].copy()
    if window.empty:
        return pd.DataFrame(
            columns=[
                "kpa_code",
                "kpa_name",
                "count_evidence",
                "avg_rating",
                "max_rating",
                "ratings_labels",
                "sample_impact",
            ]
        )

    # Convert rating to numeric when possible
    if "rating" in window.columns:
        window["rating_num"] = pd.to_numeric(window["rating"], errors="coerce")
    else:
        window["rating_num"] = None

    def _collapse(group: pd.DataFrame) -> pd.Series:
        count_evidence = len(group)
        avg_rating = group["rating_num"].dropna().mean()
        max_rating = group["rating_num"].dropna().max()
        labels = sorted({str(x) for x in group.get("rating_label", pd.Series([])).dropna().unique()})
        sample_impacts = [str(x) for x in group.get("impact_summary", pd.Series([])).dropna().head(3)]
        return pd.Series(
            {
                "count_evidence": count_evidence,
                "avg_rating": float(avg_rating) if pd.notna(avg_rating) else None,
                "max_rating": float(max_rating) if pd.notna(max_rating) else None,
                "ratings_labels": ", ".join(labels),
                "sample_impact": " | ".join(sample_impacts),
            }
        )

    grouped = (
        window.groupby(["kpa_code", "kpa_name"], dropna=False)
        .apply(_collapse)
        .reset_index()
        .sort_values(["kpa_code", "kpa_name"])
    )

    return grouped


def generate_initial_pa(profile: StaffProfile, out_dir: Path) -> Path:
    """
    Generate the Initial Performance Agreement Excel from the contract only,
    filtering out KPIs with no time/weight allocation or inactive.
    """
    rows = _filter_live_kpis(_rows_for_profile(profile))

    wb = Workbook()
    ws = wb.active
    ws.title = "Initial PA"

    headers = ["KPA Name", "Outputs", "KPI", "Weight", "Hours", "Outcomes", "Active"]
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Handle both Path and string for out_dir
    if isinstance(out_dir, str):
        out_dir = Path(out_dir)
    
    out_path = out_dir / f"PA_{profile.staff_id}_{profile.cycle_year}_initial.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return str(out_path)


def generate_mid_year_review(profile: StaffProfile, out_dir: Path) -> Path:
    """
    Generate a Mid-Year review workbook that aggregates evidence
    from months 1–6 and summarises by KPA, plus a raw evidence sheet.
    """
    df = _load_evidence(profile.staff_id, profile.cycle_year)
    mid_df = df[(df.get("month_int", 0) >= 1) & (df.get("month_int", 0) <= 6)].copy()

    wb = Workbook()

    # Sheet 1: KPA summary
    ws_summary = wb.active
    ws_summary.title = "MidYear_KPA_Summary"

    ws_summary.append(
        [
            "KPA Code",
            "KPA Name",
            "Evidence Count",
            "Average Rating",
            "Max Rating",
            "Rating Labels",
            "Sample Impact (up to 3)",
        ]
    )

    agg = _aggregate_kpa_window(df, 1, 6)
    for _, row in agg.iterrows():
        ws_summary.append(
            [
                row.get("kpa_code", ""),
                row.get("kpa_name", ""),
                row.get("count_evidence", 0),
                row.get("avg_rating", ""),
                row.get("max_rating", ""),
                row.get("ratings_labels", ""),
                row.get("sample_impact", ""),
            ]
        )

    # Sheet 2: raw evidence for months 1–6
    ws_ev = wb.create_sheet("MidYear_Evidence")
    if not mid_df.empty:
        cols = list(mid_df.columns)
        ws_ev.append(cols)
        for _, r in mid_df.iterrows():
            ws_ev.append([r.get(c, "") for c in cols])
    else:
        ws_ev.append(["No evidence found for months 1–6."])

    out_path = out_dir / f"PA_{profile.staff_id}_{profile.cycle_year}_mid_year.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def generate_final_review(profile: StaffProfile, out_dir: Path) -> Path:
    """
    Generate a Final review workbook that aggregates evidence
    from months 7–12 and summarises by KPA, plus a raw evidence sheet.
    """
    df = _load_evidence(profile.staff_id, profile.cycle_year)
    final_df = df[(df.get("month_int", 0) >= 7) & (df.get("month_int", 0) <= 12)].copy()

    wb = Workbook()

    # Sheet 1: KPA summary
    ws_summary = wb.active
    ws_summary.title = "Final_KPA_Summary"

    ws_summary.append(
        [
            "KPA Code",
            "KPA Name",
            "Evidence Count",
            "Average Rating",
            "Max Rating",
            "Rating Labels",
            "Sample Impact (up to 3)",
        ]
    )

    agg = _aggregate_kpa_window(df, 7, 12)
    for _, row in agg.iterrows():
        ws_summary.append(
            [
                row.get("kpa_code", ""),
                row.get("kpa_name", ""),
                row.get("count_evidence", 0),
                row.get("avg_rating", ""),
                row.get("max_rating", ""),
                row.get("ratings_labels", ""),
                row.get("sample_impact", ""),
            ]
        )

    # Sheet 2: raw evidence for months 7–12
    ws_ev = wb.create_sheet("Final_Evidence")
    if not final_df.empty:
        cols = list(final_df.columns)
        ws_ev.append(cols)
        for _, r in final_df.iterrows():
            ws_ev.append([r.get(c, "") for c in cols])
    else:
        ws_ev.append(["No evidence found for months 7–12."])

    out_path = out_dir / f"PA_{profile.staff_id}_{profile.cycle_year}_final.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
