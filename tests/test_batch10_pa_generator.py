from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.batch10_pa_generator import (
    Batch10Input,
    Batch10Metadata,
    Batch10Results,
    HEADERS,
    KPA_ORDER,
    generate_pa_report,
)
from backend.batch8_aggregator import FinalPerformance, KPASummary
from backend.contracts.contract_builder import MergedKPA, PerformanceContract


@pytest.fixture()
def sample_contract() -> PerformanceContract:
    kpas = {
        "KPA3": MergedKPA(
            code="KPA3",
            name="Personal Research, Innovation and/or Creative Outputs",
            hours=120.0,
            weight_pct=40.0,
            outputs=["Publish in accredited journals", "Present at conferences"],
            kpis=[
                {
                    "description": "Submit journal articles",
                    "measure": "Articles submitted",
                    "target": "2 per year",
                }
            ],
            outcomes=["Strengthen research profile"],
            active=True,
        ),
        "KPA2": MergedKPA(
            code="KPA2",
            name="Teaching and Learning, including Higher Degree Supervision",
            hours=100.0,
            weight_pct=30.0,
            outputs="Deliver undergraduate modules",
            kpis=["Complete curriculum review"],
            outcomes=[],
            active=True,
        ),
        "KPA4": MergedKPA(
            code="KPA4",
            name="Academic Leadership, Management and Administration",
            hours=60.0,
            weight_pct=20.0,
            outputs=[],
            kpis=[],
            outcomes=None,
            active=True,
        ),
        "KPA5": MergedKPA(
            code="KPA5",
            name="Social Responsiveness and Industry Involvement",
            hours=10.0,
            weight_pct=5.0,
            outputs=["Community outreach"],
            kpis=[{"description": "Host workshops", "measure": "Workshops", "target": "3"}],
            outcomes=["Engaged scholarship"],
            active=True,
        ),
        "KPA1": MergedKPA(
            code="KPA1",
            name="OHS (Occupational Health and Safety)",
            hours=5.0,
            weight_pct=3.0,
            outputs=[],
            kpis=[],
            outcomes=None,
            active=True,
        ),
        "KPA6": MergedKPA(
            code="KPA6",
            name="People Management",
            hours=5.0,
            weight_pct=2.0,
            outputs=[],
            kpis=[],
            outcomes=None,
            active=False,
        ),
    }
    return PerformanceContract(
        staff_id="55291597",
        cycle_year="2025",
        kpas=kpas,
        total_weight_pct=100.0,
        valid=True,
    )


@pytest.fixture()
def batch10_input(sample_contract: PerformanceContract) -> Batch10Input:
    metadata = Batch10Metadata(
        staff_no="55291597",
        full_name="Test User",
        year=2025,
        faculty="FEDU",
        post_level="PL5",
    )
    batch8_results = Batch10Results(
        kpa_summaries=[
            KPASummary(
                kpa_code="KPA3",
                kpa_name="Personal Research, Innovation and/or Creative Outputs",
                weight_pct=40.0,
                kcr=0.8,
                status="ACHIEVED",
                contributing_artefacts=2,
            )
        ],
        final_performance=FinalPerformance(
            overall_score=0.75,
            final_rating=4,
            final_tier="Transformational",
            justification="Deterministic",
        ),
    )
    return Batch10Input(contract=sample_contract, batch8_results=batch8_results, metadata=metadata)


def test_headers_and_ordering(batch10_input: Batch10Input, tmp_path: Path) -> None:
    out_path = generate_pa_report(batch10_input, tmp_path)
    wb = load_workbook(out_path)
    ws = wb["pa-report"]

    assert [cell.value for cell in ws[2]] == HEADERS

    kpa_names = [ws.cell(row=i, column=1).value for i in range(3, 3 + len(KPA_ORDER))]
    assert kpa_names == KPA_ORDER


def test_outputs_kpis_and_no_scores(batch10_input: Batch10Input, tmp_path: Path) -> None:
    out_path = generate_pa_report(batch10_input, tmp_path)
    wb = load_workbook(out_path)
    ws = wb["pa-report"]

    outputs_cell = ws.cell(row=3, column=2).value
    assert outputs_cell == "Publish in accredited journals\nPresent at conferences"

    kpi_cell = ws.cell(row=3, column=3).value
    assert "• Submit journal articles" in kpi_cell
    assert "Measure: Articles submitted" in kpi_cell
    assert "Target: 2 per year" in kpi_cell

    sheet_text = "\n".join(
        str(ws.cell(row=i, column=j).value or "") for i in range(1, ws.max_row + 1) for j in range(1, ws.max_column + 1)
    )
    assert "rating" not in sheet_text.lower()
    assert "tier" not in sheet_text.lower()


def test_weight_and_hours_preserved(batch10_input: Batch10Input, tmp_path: Path) -> None:
    out_path = generate_pa_report(batch10_input, tmp_path)
    wb = load_workbook(out_path)
    ws = wb["pa-report"]

    weights = [ws.cell(row=i, column=4).value for i in range(3, 9)]
    hours = [ws.cell(row=i, column=5).value for i in range(3, 9)]

    assert pytest.approx(sum(weights), rel=0.01) == 100.0
    assert hours[0] == 120.0
    assert hours[1] == 100.0
    assert hours[-1] == 5.0


def test_missing_kpa_validation(batch10_input: Batch10Input, tmp_path: Path) -> None:
    batch10_input.contract.kpas.pop("KPA6")
    with pytest.raises(ValueError):
        generate_pa_report(batch10_input, tmp_path)

