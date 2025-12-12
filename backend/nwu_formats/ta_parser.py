from __future__ import annotations

"""Parser for the NWU Task Agreement (TA) sheet."""

import json
import re
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Dict, List, Tuple, Any

from openpyxl import load_workbook  # type: ignore


@dataclass
class PerformanceKPA:
    """KPA summary extracted from a Task Agreement."""

    code: str
    name: str
    hours: float
    weight_pct: float
    outputs: List = None
    kpis: List = None
    outcomes: List = None
    active: bool = True
    context: Dict[str, Any] | None = None
    source_sheet: str | None = None
    row_number: int | None = None
    status: str = "OK"
    validation_note: str | None = None

    def __post_init__(self) -> None:
        self.outputs = [] if self.outputs is None else self.outputs
        self.kpis = [] if self.kpis is None else self.kpis
        self.outcomes = [] if self.outcomes is None else self.outcomes


@dataclass
class PerformanceContract:
    """Lightweight performance contract derived from a TA."""

    staff_id: str
    cycle_year: str
    kpas: Dict[str, PerformanceKPA]
    total_weight_pct: float
    valid: bool
    status: str = "OK"
    validation_errors: List[str] = field(default_factory=list)
    snapshot: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict:
        return {
            "staff_id": self.staff_id,
            "cycle_year": self.cycle_year,
            "total_weight_pct": self.total_weight_pct,
            "valid": self.valid,
            "kpas": {code: asdict(kpa) for code, kpa in self.kpas.items()},
            "status": self.status,
            "validation_errors": list(self.validation_errors),
            "snapshot": self.snapshot,
        }


SECTION_KPA_MAP: Dict[int, Tuple[str, str]] = {
    1: ("KPA1", "Personal Research, Innovation and/or Creative Outputs"),
    2: ("KPA2", "Teaching and Learning, including Higher Degree Supervision"),
    3: ("KPA3", "Academic Leadership, Management and Administration"),
    4: ("KPA4", "Social Responsiveness and Industry Involvement"),
    5: ("KPA5", "OHS"),
    6: ("KPA6", "People Management"),
}


def _safe_float(value) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def _extract_section_number(text: str) -> int | None:
    match = re.search(r"section\s*(\d+)", str(text), re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


def _detect_sheet(path_to_xlsx: str):
    wb = load_workbook(filename=path_to_xlsx, data_only=True)
    chosen = wb.sheetnames[0]
    task_pattern = re.compile(r"task\s*agreement", re.IGNORECASE)

    for sheet_name in wb.sheetnames:
        if task_pattern.search(sheet_name):
            chosen = sheet_name
            break
        if "task" in sheet_name.lower() and "agreement" in sheet_name.lower():
            chosen = sheet_name
            break

    return wb[chosen], wb, chosen


def _is_numeric(value: object) -> bool:
    try:
        float(value)
        return True
    except Exception:
        return False


def _build_kpa_context_from_summary(summary: Dict[str, Any], kpa_code: str) -> Dict[str, Any]:
    """Map expectation_engine TA summary fields into a KPA context payload."""

    context: Dict[str, Any] = {}
    kpa_summary = summary.get("kpa_summary", {}) or {}
    if kpa_code in kpa_summary and isinstance(kpa_summary[kpa_code], dict):
        context.update(kpa_summary[kpa_code])

    norm_hours = summary.get("norm_hours")
    if norm_hours:
        context.setdefault("norm_hours", norm_hours)

    category_map: Dict[str, List[str]] = {
        "KPA1": ["teaching", "supervision", "teaching_practice_windows"],
        "KPA2": ["ohs"],
        "KPA3": ["research"],
        "KPA4": ["leadership"],
        "KPA5": ["social"],
    }

    for key in category_map.get(kpa_code, []):
        value = summary.get(key)
        if value:
            context[key] = value

    if kpa_code == "KPA4":
        people_management = summary.get("people_management")
        if people_management:
            context["people_management_items"] = list(people_management)

    return context


def _fold_people_management_summary(summary: Dict[str, Any], director_level: bool) -> Dict[str, Any]:
    """Merge People Management metrics into KPA4 for non-director staff."""

    if director_level:
        return summary or {}

    merged_summary = dict(summary or {})
    kpa_summary = dict(merged_summary.get("kpa_summary") or {})
    people_management = merged_summary.get("people_management")

    pm_block = kpa_summary.pop("KPA6", None)
    if pm_block:
        kpa4 = dict(kpa_summary.get("KPA4") or {})
        kpa4_hours = _safe_float(kpa4.get("hours")) + _safe_float(pm_block.get("hours"))
        kpa4_weight = _safe_float(kpa4.get("weight_pct")) + _safe_float(pm_block.get("weight_pct"))
        if not kpa4.get("name"):
            kpa4["name"] = pm_block.get("name", "Academic Leadership and Management")
        kpa4["hours"] = kpa4_hours
        kpa4["weight_pct"] = kpa4_weight
        kpa_summary["KPA4"] = kpa4

    merged_summary["kpa_summary"] = kpa_summary
    if people_management:
        merged_summary["people_management"] = people_management
    return merged_summary


def parse_nwu_ta(path_to_xlsx: str, director_level: bool = False) -> PerformanceContract:
    """Parse NWU Task Agreement grand totals into a PerformanceContract."""
    validation_errors: List[str] = []
    snapshot_rows: List[Dict[str, Any]] = []
    section_totals: List[Dict[str, Any]] = []
    kpas: Dict[str, PerformanceKPA] = {}
    sheet_name = "unknown"
    wb = None

    try:
        ws, wb, sheet_name = _detect_sheet(path_to_xlsx)
    except Exception as exc:
        contract = PerformanceContract(
            staff_id="unknown_staff",
            cycle_year="unknown_year",
            kpas={},
            total_weight_pct=0.0,
            valid=False,
            status="INVALID_TA",
            validation_errors=[str(exc)],
        )
        _save_snapshot(contract)
        return contract

    try:
        total_pattern = re.compile(
            r"grand\s*total\s*[:\-]?\s*section\s*(\d+)", re.IGNORECASE
        )

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row or not row[0]:
                continue
            label = str(row[0]).strip()
            m = total_pattern.search(label)
            if not m:
                continue

            section_num = _extract_section_number(label) or _safe_float(m.group(1))
            if not section_num:
                validation_errors.append(f"Row {idx}: could not detect section number")
                continue
            section_num = int(section_num)

            mapping = SECTION_KPA_MAP.get(section_num)
            if mapping is None:
                validation_errors.append(f"Row {idx}: unknown section {section_num}")
                continue

            if len(row) <= 4:
                validation_errors.append(f"Row {idx}: missing hours/weight columns D/E")
                continue

            hours_raw = row[3]
            weight_raw = row[4] if len(row) > 4 else None
            if not _is_numeric(hours_raw) or not _is_numeric(weight_raw):
                validation_errors.append(
                    f"Row {idx}: non-numeric hours/weight in columns D/E"
                )
                continue

            hours = _safe_float(hours_raw)
            weight_pct = _safe_float(weight_raw)

            code, name = mapping
            kpa = PerformanceKPA(
                code=code,
                name=name,
                hours=hours,
                weight_pct=weight_pct,
                outputs=[],
                kpis=[],
                outcomes=[],
                active=True,
                context={},
                source_sheet=sheet_name,
                row_number=idx,
            )
            kpas[kpa.code] = kpa
            row_data = {
                "section": section_num,
                "sheet": sheet_name,
                "row_number": idx,
                "label": label,
                "hours_raw": hours_raw,
                "weight_raw": weight_raw,
                "hours": hours,
                "weight_pct": weight_pct,
            }
            snapshot_rows.append(row_data)
            section_totals.append(row_data)

        # Attach teaching modules from Addendum B if available
        modules: List[str] = []
        summary: Dict[str, Any] = {}
        try:
            from backend.expectation_engine import parse_task_agreement  # type: ignore

            summary = _fold_people_management_summary(
                parse_task_agreement(path_to_xlsx, director_level=director_level),
                director_level,
            )
            modules = summary.get("teaching_modules") or []
        except Exception:
            modules = []
            summary = {}
    except Exception as exc:
        validation_errors.append(f"Unexpected parsing error: {exc}")
        modules = []
        summary = {}

    if not section_totals:
        validation_errors.append("No GRAND TOTAL section rows found in sheet")

    if modules and "KPA2" in kpas:
        kpa2 = kpas["KPA2"]
        kpa2.context = kpa2.context or {}
        kpa2.context["modules"] = modules

    if not director_level and "KPA6" in kpas:
        pm_kpa = kpas.pop("KPA6")
        target = kpas.get("KPA4")
        if target:
            target.hours = (target.hours or 0.0) + (pm_kpa.hours or 0.0)
            target.weight_pct = (target.weight_pct or 0.0) + (pm_kpa.weight_pct or 0.0)
            pm_context_raw = getattr(pm_kpa, "context", {}) or {}
            pm_context = pm_context_raw if isinstance(pm_context_raw, dict) else {}
            if pm_context:
                combined_ctx = dict(pm_context)
                target_context_raw = getattr(target, "context", {}) or {}
                if isinstance(target_context_raw, dict):
                    combined_ctx.update(target_context_raw)
                target.context = combined_ctx
        else:
            kpas["KPA4"] = pm_kpa
            pm_kpa.code = "KPA4"
            pm_kpa.name = "Academic Leadership and Management"

    # Attach TA context buckets when available
    if summary:
        for code, kpa in kpas.items():
            context = getattr(kpa, "context", {}) or {}
            merged = dict(context)
            merged.update(_build_kpa_context_from_summary(summary, code))
            kpa.context = merged

    total_weight = sum(kpa.weight_pct for kpa in kpas.values())
    total_hours = sum(kpa.hours for kpa in kpas.values()) if kpas else 0.0
    if abs(total_weight - 100.0) >= 0.5:
        validation_errors.append(
            f"Section weights total {total_weight:.2f}, expected approximately 100"
        )
    is_valid = bool(kpas) and abs(total_weight - 100.0) < 0.5 and not validation_errors
    status = "OK" if is_valid else "INVALID_TA"

    # Fallback metadata if we cannot read it from the workbook properties
    try:
        staff_id = getattr(wb.properties, "creator", None) or "unknown_staff"
        created = getattr(wb.properties, "created", None)
        cycle_year = created.year if created else None
    except Exception:
        staff_id = "unknown_staff"
        cycle_year = None
    cycle_year = str(cycle_year) if cycle_year else "unknown_year"

    contract = PerformanceContract(
        staff_id=str(staff_id),
        cycle_year=cycle_year,
        kpas=kpas,
        total_weight_pct=total_weight,
        valid=is_valid,
        status=status,
        validation_errors=validation_errors,
        snapshot={
            "sheet": sheet_name,
            "grand_totals": snapshot_rows,
            "section_totals": section_totals,
            "modules": modules,
            "ta_parse_report": summary.get("ta_parse_report", {}),
            "norm_hours": summary.get("norm_hours"),
            "total_hours": total_hours,
            "ta_warnings": summary.get("warnings", []),
        },
    )

    _save_snapshot(contract)
    return contract


def _save_snapshot(contract: PerformanceContract) -> Path:
    base_dir = Path(__file__).resolve().parents[1] / "data" / "contracts"
    base_dir.mkdir(parents=True, exist_ok=True)

    safe_staff = (
        str(contract.staff_id).replace("/", "-").replace("\\", "-") or "unknown_staff"
    )
    safe_year = str(contract.cycle_year) or "unknown_year"
    out_path = base_dir / f"{safe_staff}_{safe_year}_TA.json"

    out_path.write_text(json.dumps(contract.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8")
    return out_path
