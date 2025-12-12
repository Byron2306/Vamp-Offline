from __future__ import annotations

"""Batch 10 – Deterministic PA Excel export.

This module renders the canonical :class:`PerformanceContract` into the official
NWU Performance Agreement Excel layout (``pa-report`` sheet) without adding any
AI-generated content or scoring fields.
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from backend.batch8_aggregator import FinalPerformance, KPASummary
from backend.contracts.contract_builder import MergedKPA, PerformanceContract


KPA_ORDER: List[str] = [
    "Personal Research, Innovation and/or Creative Outputs",
    "Teaching and Learning, including Higher Degree Supervision",
    "Academic Leadership, Management and Administration",
    "Social Responsiveness and Industry Involvement",
    "OHS (Occupational Health and Safety)",
    "People Management",
]

HEADERS: List[str] = ["KPA Name", "Outputs", "KPIs", "Weight", "Hours", "Outcomes", "Active"]

DEFAULT_PLACEHOLDER = "Not Available"


@dataclass
class Batch10Metadata:
    staff_no: str
    full_name: str
    year: int
    faculty: str
    post_level: str


@dataclass
class Batch10Results:
    kpa_summaries: Sequence[KPASummary]
    final_performance: FinalPerformance


@dataclass
class Batch10Input:
    contract: PerformanceContract
    batch8_results: Batch10Results
    metadata: Batch10Metadata


def _normalise_key(name: str) -> str:
    return " ".join(name.lower().split())


def _kpa_lookup(contract: PerformanceContract) -> Dict[str, MergedKPA]:
    lookup: Dict[str, MergedKPA] = {}
    for kpa in contract.kpas.values():
        lookup[_normalise_key(kpa.name)] = kpa
    return lookup


def _render_outputs(raw_outputs: object) -> str:
    if raw_outputs is None:
        return ""
    if isinstance(raw_outputs, str):
        return raw_outputs.strip()
    if isinstance(raw_outputs, Iterable):
        parts = [str(item).strip() for item in raw_outputs if str(item).strip()]
        return "\n".join(parts)
    return str(raw_outputs).strip()


def _extract_kpi_fields(raw: object) -> Optional[Dict[str, str]]:
    if raw is None:
        return None
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return None
        return {"text": text, "measure": DEFAULT_PLACEHOLDER, "target": DEFAULT_PLACEHOLDER}

    if isinstance(raw, dict):
        text = str(
            raw.get("description")
            or raw.get("text")
            or raw.get("kpi")
            or raw.get("kpi_text")
            or ""
        ).strip()
        if not text:
            return None
        measure = str(raw.get("measure", "")).strip() or DEFAULT_PLACEHOLDER
        target = str(raw.get("target", "")).strip() or DEFAULT_PLACEHOLDER
        return {"text": text, "measure": measure, "target": target}

    return {"text": str(raw).strip(), "measure": DEFAULT_PLACEHOLDER, "target": DEFAULT_PLACEHOLDER}


def _render_kpis(raw_kpis: object) -> str:
    if raw_kpis is None:
        return ""
    items: List[Dict[str, str]] = []
    if isinstance(raw_kpis, str):
        maybe = _extract_kpi_fields(raw_kpis)
        if maybe:
            items.append(maybe)
    elif isinstance(raw_kpis, dict):
        maybe = _extract_kpi_fields(raw_kpis)
        if maybe:
            items.append(maybe)
    elif isinstance(raw_kpis, Iterable):
        for entry in raw_kpis:
            maybe = _extract_kpi_fields(entry)
            if maybe:
                items.append(maybe)
    else:
        maybe = _extract_kpi_fields(raw_kpis)
        if maybe:
            items.append(maybe)

    blocks: List[str] = []
    for item in items:
        text = item.get("text", "").strip()
        if not text:
            continue
        measure = item.get("measure", DEFAULT_PLACEHOLDER) or DEFAULT_PLACEHOLDER
        target = item.get("target", DEFAULT_PLACEHOLDER) or DEFAULT_PLACEHOLDER
        block = f"• {text}\n  Measure: {measure}\n  Target: {target}"
        blocks.append(block)

    return "\n\n".join(blocks)


def _render_outcomes(raw_outcomes: object) -> str:
    if raw_outcomes is None:
        return "To be evaluated at year-end"
    if isinstance(raw_outcomes, str):
        return raw_outcomes.strip() or "To be evaluated at year-end"
    if isinstance(raw_outcomes, Iterable):
        parts = [str(item).strip() for item in raw_outcomes if str(item).strip()]
        if parts:
            return "\n".join(parts)
        return "To be evaluated at year-end"
    return str(raw_outcomes).strip() or "To be evaluated at year-end"


def _validate_rows(rows: List[List[object]]) -> None:
    if len(rows) < len(KPA_ORDER):
        raise ValueError("PA export aborted: missing KPA rows")

    weight_total = 0.0
    for row in rows:
        try:
            weight_total += float(row[3] or 0)
        except Exception:
            continue

    if abs(weight_total - 100.0) > 0.5:
        raise ValueError("PA export aborted: weights do not sum to ~100%")


def _apply_layout(ws) -> None:
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 8

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=3, max_col=7):
        for cell in row:
            if cell.column_letter in {"B", "C", "F"}:
                cell.alignment = wrap_alignment

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    for cell in ws[2]:
        cell.font = bold_font


def generate_pa_report(batch10_input: Batch10Input, output_dir: Path) -> Path:
    contract = batch10_input.contract
    metadata = batch10_input.metadata
    lookup = _kpa_lookup(contract)

    rows: List[List[object]] = []
    for kpa_name in KPA_ORDER:
        kpa_key = _normalise_key(kpa_name)
        kpa = lookup.get(kpa_key)

        outputs = _render_outputs(kpa.outputs if kpa else None)
        kpis = _render_kpis(kpa.kpis if kpa else None)
        weight = float(kpa.weight_pct) if kpa else 0.0
        hours = float(kpa.hours) if kpa else 0.0
        outcomes = _render_outcomes(kpa.outcomes if kpa else None)
        active_flag = "Y" if (kpa.active if kpa else True) else "N"

        rows.append([kpa_name, outputs, kpis, weight, hours, outcomes, active_flag])

    _validate_rows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "pa-report"

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Performance Agreement {metadata.staff_no} {metadata.year}"
    title_cell.font = Font(bold=True)

    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    _apply_layout(ws)

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"PA_{metadata.staff_no}_{metadata.year}.xlsx"
    wb.save(out_path)
    return out_path

