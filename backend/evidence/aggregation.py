from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook  # type: ignore

from backend.contracts.models import StaffProfile, KPA_DEFINITIONS
from backend.evidence.evidence_store import evidence_csv_path

BASE_DIR = Path(__file__).resolve().parents[1]  # backend/
DATA_DIR = BASE_DIR / "data"
TEMPLATES_DIR = DATA_DIR / "templates"
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

MID_YEAR_TEMPLATE = TEMPLATES_DIR / "mid_year_review_blank.xlsx"
FINAL_REVIEW_TEMPLATE = TEMPLATES_DIR / "final_review_blank.xlsx"

MID_YEAR_OUTPUT_DIR = DATA_DIR / "mid_year_reviews"
FINAL_OUTPUT_DIR = DATA_DIR / "final_reviews"
MID_YEAR_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
FINAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class KpaAggregate:
    kpa_code: str
    kpa_name: str
    count: int = 0
    ratings: List[float] = field(default_factory=list)
    impact_snippets: List[str] = field(default_factory=list)
    gaps_snippets: List[str] = field(default_factory=list)

    @property
    def avg_rating(self) -> Optional[float]:
        if not self.ratings:
            return None
        return sum(self.ratings) / len(self.ratings)

    @property
    def impact_summary(self) -> str:
        if not self.impact_snippets:
            return ""
        snippets = self.impact_snippets[:5]
        return " ".join(s.strip() for s in snippets if s.strip())

    @property
    def gaps_summary(self) -> str:
        if not self.gaps_snippets:
            return ""
        snippets = self.gaps_snippets[:5]
        return " ".join(s.strip() for s in snippets if s.strip())


def _ensure_mid_year_template() -> None:
    if MID_YEAR_TEMPLATE.is_file():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Mid-Year Review"
    ws.append([
        "KPA Name",
        "Evidence count (Jan–Jun)",
        "Average rating",
        "Impact highlights",
        "Risks / gaps",
    ])
    wb.save(MID_YEAR_TEMPLATE)


def _ensure_final_template() -> None:
    if FINAL_REVIEW_TEMPLATE.is_file():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Review"
    ws.append([
        "KPA Name",
        "Evidence count (Jan–Oct)",
        "Average rating (Jan–Oct)",
        "Impact highlights (full year)",
        "Risks / gaps (full year)",
    ])
    wb.save(FINAL_REVIEW_TEMPLATE)


def _load_evidence_rows(staff_id: str, cycle_year: int) -> List[Dict[str, Any]]:
    path = evidence_csv_path(staff_id, cycle_year)
    if not path.is_file():
        return []

    rows: List[Dict[str, Any]] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
    return rows


def _parse_month_bucket(value: str, default_year: int) -> Tuple[int, int]:
    if not value:
        return default_year, 0
    try:
        year_str, month_str = value.split("-", 1)
        year = int(year_str)
        month = int(month_str)
        return year, month
    except Exception:
        return default_year, 0


def _aggregate_by_kpa(
    rows: List[Dict[str, Any]],
    cycle_year: int,
    max_month: int,
) -> Dict[str, KpaAggregate]:
    agg: Dict[str, KpaAggregate] = {}

    for row in rows:
        mb = str(row.get("month_bucket", ""))
        year, month = _parse_month_bucket(mb, cycle_year)
        if year != cycle_year or month == 0 or month > max_month:
            continue

        kpa_code = (row.get("kpa_code") or "").strip() or "UNKNOWN"
        kpa_name = (row.get("kpa_name") or "").strip() or KPA_DEFINITIONS.get(kpa_code, kpa_code)

        key = kpa_code
        if key not in agg:
            agg[key] = KpaAggregate(kpa_code=kpa_code, kpa_name=kpa_name)

        item = agg[key]
        item.count += 1

        rating_raw = row.get("rating", "")
        try:
            rating = float(rating_raw)
            item.ratings.append(rating)
        except Exception:
            pass

        impact = (row.get("impact_summary") or "").strip()
        if impact:
            item.impact_snippets.append(impact)

        gaps = (row.get("risks_or_gaps") or "").strip()
        if gaps:
            item.gaps_snippets.append(gaps)

    return agg


def generate_mid_year_review_excel(profile: StaffProfile, max_month: int = 6) -> Path:
    _ensure_mid_year_template()

    rows = _load_evidence_rows(profile.staff_id, profile.cycle_year)
    agg = _aggregate_by_kpa(rows, profile.cycle_year, max_month)

    wb = load_workbook(str(MID_YEAR_TEMPLATE))
    ws = wb.active

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for kpa in profile.kpas:
        code = kpa.code
        name = kpa.name or KPA_DEFINITIONS.get(code, code)
        summary = agg.get(code) or KpaAggregate(kpa_code=code, kpa_name=name)

        avg = summary.avg_rating
        avg_str = "" if avg is None else f"{avg:.2f}"

        ws.append([
            summary.kpa_name,
            summary.count,
            avg_str,
            summary.impact_summary,
            summary.gaps_summary,
        ])

    safe_id = profile.staff_id.replace("/", "-").replace("\\", "-")
    filename = f"MidYear_{safe_id}_{profile.cycle_year}.xlsx"
    out_path = MID_YEAR_OUTPUT_DIR / filename
    wb.save(out_path)
    return out_path


def generate_final_review_excel(profile: StaffProfile, max_month: int = 10) -> Path:
    _ensure_final_template()

    rows = _load_evidence_rows(profile.staff_id, profile.cycle_year)
    agg = _aggregate_by_kpa(rows, profile.cycle_year, max_month)

    wb = load_workbook(str(FINAL_REVIEW_TEMPLATE))
    ws = wb.active

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for kpa in profile.kpas:
        code = kpa.code
        name = kpa.name or KPA_DEFINITIONS.get(code, code)
        summary = agg.get(code) or KpaAggregate(kpa_code=code, kpa_name=name)

        avg = summary.avg_rating
        avg_str = "" if avg is None else f"{avg:.2f}"

        ws.append([
            summary.kpa_name,
            summary.count,
            avg_str,
            summary.impact_summary,
            summary.gaps_summary,
        ])

    safe_id = profile.staff_id.replace("/", "-").replace("\\", "-")
    filename = f"FinalReview_{safe_id}_{profile.cycle_year}.xlsx"
    out_path = FINAL_OUTPUT_DIR / filename
    wb.save(out_path)
    return out_path
