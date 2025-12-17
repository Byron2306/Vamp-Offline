from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import load_workbook

from backend.contracts.pa_generator import DEFAULT_KPA_NAMES, PA_ORDER
from backend.llm.ollama_client import extract_json_object, query_ollama
from backend.staff_profile import StaffProfile


def _normalise(text: str) -> str:
    return " ".join((text or "").lower().split())


def _infer_kpa_code(name: str, idx: int) -> str:
    name_norm = _normalise(name)
    for code, label in DEFAULT_KPA_NAMES.items():
        if code.lower() in name_norm:
            return code
        if _normalise(label) in name_norm:
            return code
    if name_norm.startswith("kpa") and len(name_norm) >= 4 and name_norm[3].isdigit():
        return name_norm[:4].upper()
    if idx < len(PA_ORDER):
        return PA_ORDER[idx][0]
    return f"KPA{idx + 1}"


def _list_to_text(value: Any) -> str:
    if isinstance(value, list):
        return "\n".join(str(v).strip() for v in value if str(v).strip())
    return str(value).strip()


def load_pa_skeleton(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path)
    if "pa-report" not in wb.sheetnames:
        raise ValueError("Skeleton PA missing 'pa-report' sheet")
    ws = wb["pa-report"]

    headers = [cell.value or "" for cell in ws[2]]
    rows: List[Dict[str, Any]] = []
    for idx, cells in enumerate(ws.iter_rows(min_row=3, max_col=len(headers)), start=0):
        kpa_name = str(cells[0].value or "").strip()
        if not kpa_name:
            continue
        row: Dict[str, Any] = {headers[i]: cells[i].value for i in range(len(headers))}
        row["KPA"] = _infer_kpa_code(kpa_name, idx)
        row["KPA Name"] = kpa_name
        rows.append(row)
    return rows


def _derive_ai_path(skeleton_path: Path) -> Path:
    stem = skeleton_path.stem
    if "_skeleton" in stem:
        stem = stem.replace("_skeleton", "_ai")
    else:
        stem = f"{stem}_ai"
    return skeleton_path.with_name(f"{stem}{skeleton_path.suffix}")


def save_pa(rows: Sequence[Dict[str, Any]], skeleton_path: Path) -> Path:
    wb = load_workbook(skeleton_path)
    if "pa-report" not in wb.sheetnames:
        raise ValueError("Skeleton PA missing 'pa-report' sheet")
    ws = wb["pa-report"]

    headers = [cell.value or "" for cell in ws[2]]
    header_index = {h: idx for idx, h in enumerate(headers)}

    for offset, row in enumerate(rows, start=0):
        excel_row = 3 + offset
        for key in ("Outputs", "KPIs", "Outcomes"):
            if key not in header_index:
                continue
            value = row.get(key, "") if isinstance(row, dict) else ""
            ws.cell(row=excel_row, column=header_index[key] + 1).value = _list_to_text(value)

    out_path = _derive_ai_path(skeleton_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _plan_to_mapping(plan: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    if not isinstance(plan, dict):
        return {}
    if "kpas" not in plan:
        return {k: v for k, v in plan.items() if isinstance(v, dict)}

    mapping: Dict[str, Dict[str, Any]] = {}
    for block in plan.get("kpas", []) or []:
        code = str(block.get("code") or block.get("kpa") or block.get("id") or "").strip().upper()
        if not code:
            continue
        mapping[code] = {
            "outputs": block.get("outputs") or [],
            "kpis": block.get("kpis") or [],
            "outcomes": block.get("outcomes") or [],
        }
    return mapping


def _build_prompt(profile: StaffProfile, rows: Sequence[Dict[str, Any]]) -> str:
    kpa_lines: List[str] = []
    for row in rows:
        kpa_lines.append(
            f"- {row.get('KPA')}: {row.get('KPA Name', '')} | hours={row.get('Hours', '')} | weight={row.get('Weight', '')}"
        )
        outcomes = _list_to_text(row.get("Outcomes", ""))
        if outcomes:
            kpa_lines.append(f"  outcomes: {outcomes}")
    summary = "\n".join(kpa_lines)

    prompt = f"""
You are enriching an existing NWU Performance Agreement skeleton. The PA already contains one row per KPA on the 'pa-report' sheet.
Only enrich text fields; never add or remove rows and never change hours/weight.

Staff ID: {profile.staff_id}
Year: {profile.cycle_year}
KPA summary:
{summary}

Rules:
- Keep the existing KPA codes exactly as provided.
- Provide concise Outputs, measurable KPIs, and optional refined Outcomes for each KPA.
- Do NOT create new KPAs, KPIs, or alter hours/weight allocations.
- Return JSON ONLY in this shape (no narration):
{{
  "KPA1": {{"outputs": ["..."], "kpis": ["..."], "outcomes": ["..."]}},
  "KPA2": {{"outputs": ["..."], "kpis": ["..."], "outcomes": ["..."]}}
}}
"""
    return prompt.strip()


def get_ai_json(profile: StaffProfile, skeleton_rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    prompt = _build_prompt(profile, skeleton_rows)
    raw = query_ollama(prompt, format="json")
    plan = extract_json_object(raw)
    if plan == "AI_FAILED":
        return {}
    return _plan_to_mapping(plan)


def enrich_pa_with_ai(profile: StaffProfile, skeleton_path: Path, *, log=None) -> Path:
    rows = load_pa_skeleton(skeleton_path)
    ai_plan = get_ai_json(profile, rows)

    if not ai_plan:
        if log:
            log("ℹ️ AI_FAILED – keeping skeleton unchanged.")
        return skeleton_path

    for row in rows:
        kpa_code = row.get("KPA")
        ai_block = ai_plan.get(str(kpa_code)) if kpa_code is not None else None
        if not ai_block:
            continue
        row["Outputs"] = ai_block.get("outputs", row.get("Outputs", ""))
        row["KPIs"] = ai_block.get("kpis", row.get("KPIs", ""))
        row["Outcomes"] = ai_block.get("outcomes", row.get("Outcomes", ""))

    out_path = save_pa(rows, skeleton_path)
    if log:
        log(f"✅ AI enrichment applied: {out_path}")
    return out_path
