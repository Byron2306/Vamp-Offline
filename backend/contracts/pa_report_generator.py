"""
PA Report Generator - matches PA_20172672_2025_final.xlsx format exactly.

Generates Performance Agreement with:
- KPA Name
- Outputs (bulleted list from TA)
- KPIs (from kpi_taxonomy)
- Weight (%)
- Hours
- Outcomes (from outcomes_library)
- Active (Y/N)
"""

import json
from pathlib import Path
from typing import Any, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

REPO_ROOT = Path(__file__).resolve().parents[2]

def _load_kpi_taxonomy() -> Dict[str, List[str]]:
    """Load KPI taxonomy for each KPA."""
    path = REPO_ROOT / "kpi_taxonomy_nwu_education.json"
    if path.exists():
        raw = json.loads(path.read_text(encoding="utf-8"))
        # Map from category names to KPA codes
        return {
            "KPA1": raw.get("teaching", []),
            "KPA2": raw.get("ohs", []),
            "KPA3": raw.get("research", []),
            "KPA4": raw.get("leadership", []),
            "KPA5": raw.get("social", [])
        }
    return {}


def _load_outcomes_library() -> str:
    """Load and format outcomes from outcomes_library.json."""
    path = REPO_ROOT / "nwu_values_vocabulary.json"
    if path.exists():
        values = json.loads(path.read_text(encoding="utf-8"))
        return "; ".join(values.keys())
    return "Excellence; Integrity; Innovation; Accountability; Social Responsiveness"


def _bullet_list(items: List[str]) -> str:
    """Format list items with bullet points."""
    cleaned = [str(x).strip() for x in items if str(x).strip()]
    return "\n".join([f"• {x}" for x in cleaned])


def _extract_outputs_from_contract(contract_data: Dict[str, Any], kpa_code: str) -> List[str]:
    """Extract all outputs/tasks for a KPA from contract data."""
    outputs = []
    
    if kpa_code == "KPA1":  # Teaching and Learning
        # Teaching modules
        modules = contract_data.get("teaching_modules", [])
        if isinstance(modules, list):
            outputs.extend(modules)
        
        # Teaching activities
        teaching = contract_data.get("teaching", [])
        if isinstance(teaching, list):
            outputs.extend(teaching)
        
        # Supervision
        supervision = contract_data.get("supervision", [])
        if isinstance(supervision, list):
            outputs.extend(supervision)
        
        # Practice windows
        windows = contract_data.get("teaching_practice_windows", [])
        if isinstance(windows, list):
            outputs.extend(windows)
    
    elif kpa_code == "KPA2":  # OHS
        ohs = contract_data.get("ohs", [])
        if isinstance(ohs, list) and ohs:
            outputs.extend(ohs)
        else:
            outputs.append("Compliance with institutional OHS requirements")
    
    elif kpa_code == "KPA3":  # Research
        research = contract_data.get("research", [])
        if isinstance(research, list):
            outputs.extend(research)
    
    elif kpa_code == "KPA4":  # Leadership
        leadership = contract_data.get("leadership", [])
        if isinstance(leadership, list):
            outputs.extend(leadership)
    
    elif kpa_code == "KPA5":  # Social Responsiveness
        social = contract_data.get("social", [])
        if isinstance(social, list):
            outputs.extend(social)
    
    return outputs


def generate_pa_report(contract_data: Dict[str, Any], staff_id: str, year: int) -> Dict[str, Any]:
    """
    Generate PA report data matching Excel format.
    
    Returns:
    {
        "rows": [
            {
                "kpa_name": "Teaching and Learning",
                "outputs": "• Module 1\n• Module 2",
                "kpis": "Curriculum delivery\nAssessment & feedback",
                "weight": 47.54,
                "hours": 792.99,
                "outcomes": "Excellence; Integrity; ...",
                "active": "Y"
            },
            ...
        ],
        "staff_id": str,
        "year": int
    }
    """
    kpa_summary = contract_data.get("kpa_summary", {})
    kpi_taxonomy = _load_kpi_taxonomy()
    outcomes_text = _load_outcomes_library()
    
    # KPA order matching Excel
    kpa_order = [
        ("KPA1", "Teaching and Learning"),
        ("KPA3", "Research and Innovation / Creative Outputs"),
        ("KPA5", "Social Responsiveness / Community and Industry Engagement"),
        ("KPA4", "Academic Leadership and Management"),
        ("KPA2", "Occupational Health and Safety")
    ]
    
    rows = []
    
    for kpa_code, default_name in kpa_order:
        kpa_info = kpa_summary.get(kpa_code, {})
        kpa_name = kpa_info.get("name", default_name)
        hours = kpa_info.get("hours", 0.0)
        weight = kpa_info.get("weight_pct", 0.0)
        
        # Extract outputs from contract
        outputs_list = _extract_outputs_from_contract(contract_data, kpa_code)
        outputs_text = _bullet_list(outputs_list) if outputs_list else ""
        
        # Get KPIs from taxonomy
        kpis_list = kpi_taxonomy.get(kpa_code, [])
        kpis_text = "\n".join(kpis_list[:5]) if kpis_list else ""  # Limit to 5 KPIs
        
        # OHS special case
        if kpa_code == "KPA2":
            if not outputs_text:
                outputs_text = "Compliance with institutional Occupational Health and Safety requirements"
            if not kpis_text:
                kpis_text = "Compliance with institutional Occupational Health and Safety requirements"
            if hours == 0:
                hours = 2.0
            if weight == 0:
                weight = 2.0
        
        rows.append({
            "kpa_code": kpa_code,
            "kpa_name": kpa_name,
            "outputs": outputs_text,
            "kpis": kpis_text,
            "weight": weight,
            "hours": hours,
            "outcomes": outcomes_text,
            "active": "Y"
        })
    
    return {
        "rows": rows,
        "staff_id": staff_id,
        "year": year,
        "title": f"Performance Agreement {staff_id} {year}"
    }


def export_pa_to_excel(pa_data: Dict[str, Any], output_path: Path) -> Path:
    """Export PA report data to Excel file matching the reference format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "pa-report"
    
    # Title row
    ws.append([pa_data["title"], None, None, None, None, None, None])
    
    # Header row
    ws.append(["KPA Name", "Outputs", "KPIs", "Weight", "Hours", "Outcomes", "Active"])
    
    # Data rows
    for row in pa_data["rows"]:
        ws.append([
            row["kpa_name"],
            row["outputs"],
            row["kpis"],
            row["weight"],
            row["hours"],
            row["outcomes"],
            row["active"]
        ])
    
    # Formatting
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 8
    
    # Wrap text for outputs, KPIs, outcomes
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=3, max_col=7):
        for cell in row:
            if cell.column_letter in {"B", "C", "F"}:
                cell.alignment = wrap_alignment
    
    # Bold headers
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    for cell in ws[2]:
        cell.font = bold_font
    
    wb.save(output_path)
    return output_path
