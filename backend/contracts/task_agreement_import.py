from __future__ import annotations

"""
Task Agreement importer for the offline VAMP app.

This version is deliberately aligned with backend.staff_profile so that the
Tkinter GUI, PA Excel generator and this importer all share the SAME
StaffProfile / KPA / KPI structures and JSON files.

Usage from the GUI:

    from backend.contracts.task_agreement_import import import_task_agreement_excel
    import_task_agreement_excel(self.staff_profile, Path(path))

Design rules
------------

* Only create KPIs for rows that have a realistic HOURS value (0 < hours <= 500).
  This avoids picking up section headers, totals, etc.
* KPA is inferred from nearby text in the first two columns (section labels).
* Description is taken from the main task description columns; we avoid labels
  like "Number of hours", "Hours", etc.
* We never create more than the 5 official KPAs; any unknown headings are
  mapped best-effort to KPA1–KPA5 or skipped.

If the template layout changes slightly, this should still behave sensibly:
rows that can’t be parsed into a sensible KPI are simply ignored.
"""

from pathlib import Path
from typing import Dict, List, Optional, Any

from openpyxl import load_workbook  # type: ignore

from backend.staff_profile import StaffProfile, KPA, KPI, DEFAULT_KPAS


def _norm(text: str) -> str:
    return " ".join(text.lower().strip().split())


def _guess_kpa_code(header: str) -> Optional[str]:
    """
    Map section header text to one of the 5 official KPA codes.
    """
    h = _norm(header)

    if "teaching" in h and "learning" in h:
        return "KPA1"
    if "occupational" in h and "health" in h:
        return "KPA2"
    if "ohs" in h:
        return "KPA2"
    if "personal research" in h or "research, innovation" in h or "creative outputs" in h:
        return "KPA3"
    if "academic leadership" in h or ("management" in h and "administration" in h):
        return "KPA4"
    if "social responsiveness" in h or "industry involvement" in h or "community engagement" in h:
        return "KPA5"

    # Fallback: any generic 'section KPAx' text
    if "kpa1" in h:
        return "KPA1"
    if "kpa2" in h:
        return "KPA2"
    if "kpa3" in h:
        return "KPA3"
    if "kpa4" in h:
        return "KPA4"
    if "kpa5" in h:
        return "KPA5"

    return None


def _ensure_kpa_map(profile: StaffProfile) -> Dict[str, KPA]:
    """
    Ensure the profile has the 5 canonical KPAs and return a code→KPA mapping.
    """
    kpas_by_code: Dict[str, KPA] = {k.code: k for k in profile.kpas}

    for code, name in DEFAULT_KPAS:
        if code not in kpas_by_code:
            kpas_by_code[code] = KPA(code=code, name=name, weight=None, hours=None, kpis=[])

    # Keep profile.kpas in canonical order
    profile.kpas = [kpas_by_code[code] for code, _ in DEFAULT_KPAS]
    return kpas_by_code


def _get_hours(raw: Any) -> float:
    try:
        if raw is None:
            return 0.0
        if isinstance(raw, (int, float)):
            return float(raw)
        s = str(raw).strip()
        if not s:
            return 0.0
        return float(s)
    except Exception:
        return 0.0


def _pick_description(row: List[Any]) -> str:
    """
    Choose a sensible description from the row.

    In the FEDU templates the description typically sits in the 2nd or 3rd
    column. We avoid generic labels such as "Hours" or "Number of hours".
    """
    candidates: List[str] = []
    for cell in row[1:4]:  # columns 2–4 are usually description-ish
        if isinstance(cell, str):
            txt = cell.strip()
            if not txt:
                continue
            low = txt.lower()
            if low.startswith("hours") or "number of hours" in low:
                continue
            candidates.append(txt)

    return candidates[0] if candidates else ""


def import_task_agreement_excel(profile: StaffProfile, xlsx_path: Path) -> StaffProfile:
    """
    Parse an NWU / FEDU Task Agreement Excel file and enrich the given
    StaffProfile with KPIs.

    The function modifies the profile in-place and then calls profile.save().
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Task Agreement Excel not found: {xlsx_path}")

    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb.active

    kpa_map = _ensure_kpa_map(profile)
    current_kpa_code: Optional[str] = None

    for row in ws.iter_rows(values_only=True):
        cells: List[Any] = list(row)
        if not any(cells):
            continue

        # Header / section detection from first two columns
        first = str(cells[0]) if cells[0] is not None else ""
        second = str(cells[1]) if len(cells) > 1 and cells[1] is not None else ""
        header = f"{first} {second}".strip()
        new_kpa = _guess_kpa_code(header)
        if new_kpa is not None:
            current_kpa_code = new_kpa
            continue  # section row, not a KPI row

        # We need a valid current KPA to attach the KPI
        if current_kpa_code is None:
            continue

        # Heuristic: hours is usually towards the right (e.g. column 4 or 5)
        hours_candidates = cells[3:6]
        hours = 0.0
        for hraw in hours_candidates:
            hours = _get_hours(hraw)
            if hours > 0:
                break

        # Filter out non-task rows
        if not (0.0 < hours <= 500.0):
            continue

        desc = _pick_description(cells)
        if not desc:
            continue

        # Optional weight (percentage) from the next numeric cell
        weight = 0.0
        for wraw in cells[4:7]:
            weight = _get_hours(wraw)
            if weight > 0:
                break

        kpa_obj = kpa_map.get(current_kpa_code)
        if kpa_obj is None:
            continue

        kpi = KPI(
            kpi_id=None,
            description=desc,
            outputs="",
            outcomes="",
            weight=weight if weight > 0 else None,
            hours=hours,
            active=True,
        )
        kpa_obj.kpis.append(kpi)

    # Persist contract JSON via StaffProfile.save()
    profile.save()
    return profile
