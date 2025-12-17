            from __future__ import annotations

            import json
            from pathlib import Path
            from typing import Any, Dict, List
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            DEFAULT_KPA_ORDER = [
                ("KPA1", "Teaching and Learning"),
                ("KPA2", "Research and Innovation / Creative Outputs"),
                ("KPA3", "Social Responsiveness / Community and Industry Engagement"),
                ("KPA4", "Academic Leadership and Management"),
                ("KPA5", "Occupational Health and Safety"),
            ]

            FIXED_OHS_WEIGHT = 2.0

            def _load_json(rel_path: str) -> Any:
                p = Path(__file__).resolve().parents[1] / "data" / rel_path
                return json.loads(p.read_text(encoding="utf-8"))

            def _bullet(lines: List[str]) -> str:
                cleaned = [str(x).strip() for x in lines if str(x).strip()]
                return "
".join([f"• {x}" for x in cleaned])

            def _get_ta_outputs_for_kpa(ta_context: Dict[str, Any], kpa_code: str, kpa_name: str) -> List[str]:
                kpas = ta_context.get("kpas") or ta_context.get("KPA") or {}
                if isinstance(kpas, dict):
                    block = kpas.get(kpa_code) or kpas.get(kpa_name) or {}
                    if isinstance(block, dict):
                        for key in ("outputs", "tasks", "task_list", "activities", "outcomes"):
                            v = block.get(key)
                            if isinstance(v, list):
                                return v
                            if isinstance(v, str) and v.strip():
                                return [ln.strip("• ").strip() for ln in v.splitlines() if ln.strip()]
                return []

            def _get_hours_weights(ta_context: Dict[str, Any]) -> Dict[str, Dict[str, float]]:
                out: Dict[str, Dict[str, float]] = {}
                rows = ta_context.get("kpa_breakdown") or ta_context.get("kpa_breakdown_rows") or []
                if isinstance(rows, list):
                    for r in rows:
                        if isinstance(r, dict):
                            k = r.get("kpa") or r.get("KPA")
                            if not k:
                                continue
                            out[k] = {
                                "hours": float(r.get("hours") or r.get("Hours") or 0.0),
                                "weight": float(r.get("weight") or r.get("Weight") or r.get("Weight%") or 0.0),
                            }
                return out

            def _lock_ohs_weights(breakdown: Dict[str, Dict[str, float]]) -> Dict[str, Dict[str, float]]:
                weights = {k: v.get("weight", 0.0) for k, v in breakdown.items()}
                weights["KPA5"] = FIXED_OHS_WEIGHT
                total_non_ohs = sum(w for k, w in weights.items() if k != "KPA5")
                target_non_ohs = 100.0 - FIXED_OHS_WEIGHT
                if total_non_ohs > 0:
                    scale = target_non_ohs / total_non_ohs
                    for k in list(weights.keys()):
                        if k != "KPA5":
                            weights[k] = round(weights[k] * scale, 2)
                for k, w in weights.items():
                    breakdown.setdefault(k, {})
                    breakdown[k]["weight"] = w
                return breakdown

            def build_final_pa_workbook(ta_context: Dict[str, Any], staff_no: str, year: str, out_dir: str) -> str:
                kpi_lib = _load_json("kpi_taxonomy_nwu_education.json")
                values = _load_json("nwu_values_vocabulary.json")
                outcomes_text = "; ".join(values.keys())

                breakdown = _lock_ohs_weights(_get_hours_weights(ta_context))

                wb = Workbook()
                ws = wb.active
                ws.title = "pa-report"

                ws.append([f"Performance Agreement {staff_no} {year}", None, None, None, None, None, None])
                ws.append(["KPA Name", "Outputs", "KPIs", "Weight", "Hours", "Outcomes", "Active"])

                for kpa_code, kpa_name in DEFAULT_KPA_ORDER:
                    hours = breakdown.get(kpa_code, {}).get("hours", 0.0)
                    weight = breakdown.get(kpa_code, {}).get("weight", 0.0)

                    if kpa_code == "KPA5":
                        outputs = ["Compliance with institutional Occupational Health and Safety requirements"]
                        kpis = ["Compliance with institutional Occupational Health and Safety requirements"]
                        hours = hours or 2.0
                        weight = FIXED_OHS_WEIGHT
                    else:
                        outputs = _get_ta_outputs_for_kpa(ta_context, kpa_code, kpa_name)
                        kpis_all = kpi_lib.get(kpa_code, [])
                        kpis = kpis_all[:4] if len(kpis_all) > 4 else kpis_all

                    ws.append([kpa_name, _bullet(outputs), _bullet(kpis), weight, hours, outcomes_text, "Y"])

                for col in range(1, 8):
                    max_len = 10
                    for row in range(1, ws.max_row + 1):
                        v = ws.cell(row=row, column=col).value
                        if v is None:
                            continue
                        s = str(v)
                        max_len = max(max_len, min(len(s), 80))
                    ws.column_dimensions[get_column_letter(col)].width = max_len + 2

                out_path = Path(out_dir) / f"PA_{staff_no}_{year}_final.xlsx"
                wb.save(out_path)
                return str(out_path)
